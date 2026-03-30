from __future__ import annotations

import tkinter as tk
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from tkinter import filedialog, ttk
from typing import Callable

import pandas as pd
from openpyxl import load_workbook

from ..excel_application import collect_excel_event_candidates

_SPAN_LABELS = {
    "5d": "5日",
    "3d_left": "3日前寄せ",
    "3d_center": "3日中央",
    "3d_right": "3日後寄せ",
}


@dataclass(frozen=True)
class ExcelCandidateSelection:
    event_key: str
    source_path: Path
    source_alias: str
    source_index: int
    event_date: date
    sheet_name: str
    is_resplit: bool

    def display_label(self) -> str:
        return f"{self.event_date:%Y-%m-%d} | [#{self.source_index} {self.source_alias}] {self.sheet_name}"


@dataclass
class ExcelModePanel:
    parent: tk.Misc
    frame: ttk.Frame
    span_var: tk.StringVar
    sheet_listbox: tk.Listbox
    sheet_count_var: tk.StringVar
    selected_count_var: tk.StringVar
    span_row: ttk.Frame
    input_excel_var: tk.StringVar
    source_listbox: tk.Listbox
    on_log: Callable[[str], None] | None
    _source_paths: list[Path]
    _candidates: list[ExcelCandidateSelection]

    @classmethod
    def create(
        cls,
        parent: tk.Misc,
        *,
        build_path_row: Callable[..., ttk.Frame],  # noqa: ARG003
        input_excel_var: tk.StringVar,
        on_log: Callable[[str], None] | None = None,
    ) -> "ExcelModePanel":
        frame = ttk.Frame(parent)
        span_var = tk.StringVar(value="5d")

        source_section = ttk.LabelFrame(frame, text="入力Excelファイル（複数可 / 上から処理順）", padding=6)
        source_section.pack(fill=tk.BOTH, expand=False, pady=(0, 6))
        source_section.columnconfigure(0, weight=1)
        source_section.rowconfigure(0, weight=1)

        source_list_wrap = ttk.Frame(source_section)
        source_list_wrap.grid(row=0, column=0, sticky="nsew")
        source_list_wrap.columnconfigure(0, weight=1)
        source_list_wrap.rowconfigure(0, weight=1)
        source_listbox = tk.Listbox(source_list_wrap, selectmode=tk.EXTENDED, height=4, exportselection=False)
        source_listbox.grid(row=0, column=0, sticky="nsew")
        source_scroll = ttk.Scrollbar(source_list_wrap, orient=tk.VERTICAL, command=source_listbox.yview)
        source_scroll.grid(row=0, column=1, sticky="ns")
        source_listbox.configure(yscrollcommand=source_scroll.set)

        source_buttons = ttk.Frame(source_section)
        source_buttons.grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Button(source_buttons, text="Excel追加...", command=lambda: None).pack(side=tk.LEFT)
        ttk.Button(source_buttons, text="上へ", command=lambda: None, width=5).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(source_buttons, text="下へ", command=lambda: None, width=5).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(source_buttons, text="削除", command=lambda: None, width=5).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(source_buttons, text="クリア", command=lambda: None, width=6).pack(side=tk.LEFT, padx=(6, 0))

        opt_row = ttk.Frame(frame)
        opt_row.pack(fill=tk.X, pady=(2, 4))
        opt_row.columnconfigure(1, weight=1)
        ttk.Label(opt_row, text="グラフ期間").grid(row=0, column=0, sticky="w", padx=(0, 6))
        span_wrap = ttk.Frame(opt_row)
        span_wrap.grid(row=0, column=1, sticky="w")
        ttk.Radiobutton(span_wrap, text="5日", value="5d", variable=span_var).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Radiobutton(span_wrap, text="3日前寄せ", value="3d_left", variable=span_var).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Radiobutton(span_wrap, text="3日中央", value="3d_center", variable=span_var).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Radiobutton(span_wrap, text="3日後寄せ", value="3d_right", variable=span_var).pack(side=tk.LEFT)

        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(toolbar, text="イベント候補（複数選択可）", width=22).pack(side=tk.LEFT)
        ttk.Button(toolbar, text="候補日を一括選択（再分割優先）", command=lambda: None).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(toolbar, text="候補更新", command=lambda: None, width=8).pack(side=tk.RIGHT)

        list_wrap = ttk.Frame(frame)
        list_wrap.pack(fill=tk.BOTH, expand=True, pady=(0, 2))
        sheet_listbox = tk.Listbox(list_wrap, selectmode=tk.EXTENDED, height=10, exportselection=False)
        sheet_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = ttk.Scrollbar(list_wrap, orient=tk.VERTICAL, command=sheet_listbox.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        sheet_listbox.configure(yscrollcommand=scroll.set)
        sheet_count_var = tk.StringVar(value="候補: 0件")
        selected_count_var = tk.StringVar(value="選択中: 0件")
        count_row = ttk.Frame(frame)
        count_row.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(count_row, textvariable=sheet_count_var).pack(side=tk.LEFT)
        ttk.Label(count_row, textvariable=selected_count_var).pack(side=tk.RIGHT)

        panel = cls(
            parent=parent,
            frame=frame,
            span_var=span_var,
            sheet_listbox=sheet_listbox,
            sheet_count_var=sheet_count_var,
            selected_count_var=selected_count_var,
            span_row=opt_row,
            input_excel_var=input_excel_var,
            on_log=on_log,
            source_listbox=source_listbox,
            _source_paths=[],
            _candidates=[],
        )

        toolbar_buttons = [c for c in toolbar.winfo_children() if isinstance(c, ttk.Button)]
        if len(toolbar_buttons) >= 2:
            # pack順で [一括選択, 候補更新]
            toolbar_buttons[0].configure(command=panel.select_all_dates_prefer_resplit)
            toolbar_buttons[1].configure(command=lambda p=panel: p.refresh_candidates(p.input_excel_var.get().strip()))
        source_buttons_list = [c for c in source_buttons.winfo_children() if isinstance(c, ttk.Button)]
        if len(source_buttons_list) >= 5:
            source_buttons_list[0].configure(command=panel._on_add_excels)
            source_buttons_list[1].configure(command=lambda p=panel: p._move_sources(-1))
            source_buttons_list[2].configure(command=lambda p=panel: p._move_sources(1))
            source_buttons_list[3].configure(command=panel._on_remove_sources)
            source_buttons_list[4].configure(command=panel._on_clear_excels)
        sheet_listbox.bind("<<ListboxSelect>>", lambda _e: panel._update_selected_count())
        panel.refresh_candidates(input_excel_var.get().strip())
        return panel

    def show(self, *, before: ttk.Frame | None = None) -> None:
        if before is not None:
            self.frame.pack(fill=tk.BOTH, expand=True, pady=0, before=before)
        else:
            self.frame.pack(fill=tk.BOTH, expand=True, pady=0)

    def hide(self) -> None:
        self.frame.pack_forget()

    def _on_add_excels(self) -> None:
        selected = filedialog.askopenfilenames(
            title="入力Excelファイルを選択（複数可）",
            filetypes=[("Excel", "*.xlsx;*.xls"), ("すべて", "*.*")],
        )
        if not selected:
            return
        incoming = [Path(p) for p in selected]
        existing = list(self._source_paths)
        seen = {p.resolve() if p.exists() else p for p in existing}
        for path in incoming:
            key = path.resolve() if path.exists() else path
            if key in seen:
                continue
            seen.add(key)
            existing.append(path)
        self._set_sources(existing, keep_selection=True)

    def _on_remove_sources(self) -> None:
        selected = set(self.source_listbox.curselection())
        if not selected:
            return
        kept = [path for idx, path in enumerate(self._source_paths) if idx not in selected]
        self._set_sources(kept, keep_selection=False)

    def _on_clear_excels(self) -> None:
        self._set_sources([], keep_selection=False)

    def _move_sources(self, delta: int) -> None:
        selected = sorted(self.source_listbox.curselection())
        if not selected:
            return
        sources = list(self._source_paths)
        if delta < 0:
            for idx in selected:
                if idx <= 0:
                    continue
                sources[idx - 1], sources[idx] = sources[idx], sources[idx - 1]
            new_selection = [max(0, i - 1) for i in selected]
        else:
            for idx in reversed(selected):
                if idx >= len(sources) - 1:
                    continue
                sources[idx + 1], sources[idx] = sources[idx], sources[idx + 1]
            new_selection = [min(len(sources) - 1, i + 1) for i in selected]
        self._set_sources(sources, keep_selection=True)
        self.source_listbox.selection_clear(0, tk.END)
        for idx in new_selection:
            self.source_listbox.selection_set(idx)

    def _set_sources(self, paths: list[Path], *, keep_selection: bool) -> None:
        selected_event_keys = set(self.get_selected_event_keys()) if keep_selection else set()
        self._source_paths = list(paths)
        self._sync_input_var_from_sources()
        self._render_source_list()
        self.refresh_candidates(self.input_excel_var.get().strip())
        if selected_event_keys:
            self.select_by_event_keys(selected_event_keys)

    def _sync_input_var_from_sources(self) -> None:
        self.input_excel_var.set(";".join(str(path) for path in self._source_paths))

    def _render_source_list(self) -> None:
        self.source_listbox.delete(0, tk.END)
        aliases = self._build_aliases(self._source_paths)
        for idx, path in enumerate(self._source_paths, start=1):
            alias = aliases[path]
            self.source_listbox.insert(tk.END, f"{idx}. {alias} ({path})")

    @staticmethod
    def _split_excel_paths(raw: str) -> list[Path]:
        items = [part.strip() for part in raw.replace("\n", ";").split(";") if part.strip()]
        paths: list[Path] = []
        seen: set[Path] = set()
        for item in items:
            path = Path(item)
            key = path.resolve() if path.exists() else path
            if key in seen:
                continue
            seen.add(key)
            paths.append(path)
        return paths

    @staticmethod
    def _build_aliases(paths: list[Path]) -> dict[Path, str]:
        counts: dict[str, int] = {}
        aliases: dict[Path, str] = {}
        for path in paths:
            base = path.name
            counts[base] = counts.get(base, 0) + 1
            aliases[path] = base if counts[base] == 1 else f"{base} ({counts[base]})"
        return aliases

    def get_input_excel_paths(self) -> list[Path]:
        return list(self._source_paths)

    def refresh_candidates(self, excel_paths_raw: str) -> None:
        selected_keys = set(self.get_selected_event_keys())
        self.sheet_listbox.delete(0, tk.END)
        self._candidates = []

        parsed_paths = self._split_excel_paths(excel_paths_raw)
        self._source_paths = parsed_paths
        self._render_source_list()

        if not parsed_paths:
            self.sheet_count_var.set("候補: 0件（Excel未指定）")
            self._update_selected_count()
            return

        aliases = self._build_aliases(parsed_paths)
        index_map = {path: i + 1 for i, path in enumerate(parsed_paths)}
        failure_count = 0
        for path in parsed_paths:
            alias = aliases[path]
            try:
                candidates = collect_excel_event_candidates(path)
            except Exception:  # noqa: BLE001
                failure_count += 1
                continue
            for item in candidates:
                event_key = f"{path.resolve()}::{item.sheet_name}"
                self._candidates.append(
                    ExcelCandidateSelection(
                        event_key=event_key,
                        source_path=path,
                        source_alias=alias,
                        source_index=index_map[path],
                        event_date=item.event_date,
                        sheet_name=item.sheet_name,
                        is_resplit=item.is_resplit,
                    )
                )

        self._candidates.sort(key=lambda c: (c.source_index, c.event_date, c.sheet_name))
        for item in self._candidates:
            self.sheet_listbox.insert(tk.END, item.display_label())

        source_count = len(parsed_paths) - failure_count
        self.sheet_count_var.set(
            f"候補: {len(self._candidates)}件（Excel: {source_count}件"
            + (f" / 読込失敗: {failure_count}件" if failure_count else "")
            + "）"
        )
        self.select_by_event_keys(selected_keys)

    def _update_selected_count(self) -> None:
        self.selected_count_var.set(f"選択中: {len(self.sheet_listbox.curselection())}件")

    @staticmethod
    def _preferred_keys_by_date(candidates: list[ExcelCandidateSelection]) -> tuple[set[str], int, int]:
        by_date: dict[date, list[ExcelCandidateSelection]] = defaultdict(list)
        for item in candidates:
            by_date[item.event_date].append(item)
        selected: set[str] = set()
        resplit_days = 0
        normal_days = 0
        for event_date in sorted(by_date):
            group = by_date[event_date]
            resplits = [item for item in group if item.is_resplit]
            if resplits:
                resplit_days += 1
                preferred = resplits
            else:
                normal_days += 1
                preferred = group
            selected.update(item.event_key for item in preferred)
        return selected, resplit_days, normal_days

    def select_all_dates_prefer_resplit(self) -> None:
        keys, resplit_days, normal_days = self._preferred_keys_by_date(self._candidates)
        self.select_by_event_keys(keys)
        if self.on_log is not None:
            self.on_log(
                "候補日一括選択(再分割優先): "
                f"選択{len(keys)}件 / 再分割優先{resplit_days}日 / 通常{normal_days}日"
            )

    def get_selected_events(self) -> list[ExcelCandidateSelection]:
        return [self._candidates[i] for i in self.sheet_listbox.curselection() if 0 <= i < len(self._candidates)]

    def get_selected_event_keys(self) -> list[str]:
        return [item.event_key for item in self.get_selected_events()]

    def select_by_event_keys(self, keys: set[str]) -> None:
        self.sheet_listbox.selection_clear(0, tk.END)
        if keys:
            for i, item in enumerate(self._candidates):
                if item.event_key in keys:
                    self.sheet_listbox.selection_set(i)
        self._update_selected_count()

    def get_selected_sheet_names(self) -> list[str]:
        return [item.sheet_name for item in self.get_selected_events()]

    def get_span(self) -> str:
        value = self.span_var.get().strip()
        return value if value in _SPAN_LABELS else "5d"

    def get_span_label(self) -> str:
        return _SPAN_LABELS.get(self.get_span(), "5日")

    def get_preview_span(self) -> str:
        return "5d" if self.get_span() == "5d" else "3d"

    def set_form_label_minsize(self, pixels: int) -> None:
        minsize = max(0, int(pixels))
        self.span_row.columnconfigure(0, minsize=minsize)

    def build_preview_frame(self, _excel_path_unused: str) -> pd.DataFrame | None:
        selected = self.get_selected_events()
        if not selected:
            return None
        target = selected[0]
        wb = load_workbook(target.source_path, data_only=True, read_only=True)
        try:
            if target.sheet_name not in wb.sheetnames:
                return None
            ws = wb[target.sheet_name]
            observed: list[object] = []
            rainfall: list[object] = []
            for r in ws.iter_rows(min_row=5, max_col=17, values_only=True):
                b = r[1] if len(r) > 1 else None
                q = r[16] if len(r) > 16 else None
                if b is None and q is None:
                    continue
                observed.append(b)
                rainfall.append(q)
            frame = pd.DataFrame({"observed_at": pd.to_datetime(observed, errors="coerce"), "rainfall_mm": rainfall})
            frame = frame.dropna(subset=["observed_at", "rainfall_mm"]).copy()
            frame["rainfall_mm"] = pd.to_numeric(frame["rainfall_mm"], errors="coerce")
            frame = frame.dropna(subset=["rainfall_mm"]).sort_values("observed_at").reset_index(drop=True)
            frame["observed_at"] = frame["observed_at"] - timedelta(hours=1)
            span = self.get_preview_span()
            if span == "3d" and len(frame) >= 72:
                start_idx = max(0, (len(frame) - 72) // 2)
                frame = frame.iloc[start_idx : start_idx + 72].copy()
            elif span == "5d" and len(frame) >= 120:
                start_idx = max(0, (len(frame) - 120) // 2)
                frame = frame.iloc[start_idx : start_idx + 120].copy()
            return frame
        finally:
            wb.close()
