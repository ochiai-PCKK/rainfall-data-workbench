# pyright: reportArgumentType=false, reportCallIssue=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .errors import ZipFlowError
from .graph_builder import build_metric_frame, render_region_plots_reference
from .graph_renderer_reference import compute_axis_tops, prepare_reference_window
from .logger import build_logger
from .style_profile import load_style_profile

_DATE_SHEET_RE = re.compile(r"^\d{4}\.\d{2}\.\d{2}$")
_RESPLIT_PREFIX = "【再分割】"
_REGION_KEY_EXCEL_DEFAULT = "nishiyoke_higashiyoke"
_REGION_LABEL_EXCEL_DEFAULT = "西除川+東除川"


def build_excel_filename_prefix(source_alias: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z_-]+", "_", source_alias).strip("_")
    return f"excel_{safe or 'source'}_"


@dataclass(frozen=True)
class ExcelSelectedEvent:
    source_path: Path
    source_alias: str
    sheet_name: str
    event_date: date
    is_resplit: bool


@dataclass(frozen=True)
class ExcelRunConfig:
    input_excels: tuple[Path, ...]
    output_root: Path
    selected_events: tuple[ExcelSelectedEvent, ...]
    graph_span: str  # 5d | 3d_left | 3d_center | 3d_right
    ref_graph_kinds: tuple[str, ...]  # sum | mean
    export_svg: bool
    enable_log: bool
    style_profile_path: Path | None = None
    on_conflict: str = "rename"
    region_key: str = _REGION_KEY_EXCEL_DEFAULT
    region_label: str = _REGION_LABEL_EXCEL_DEFAULT


@dataclass(frozen=True)
class ExcelEventCandidate:
    event_date: date
    sheet_name: str
    is_resplit: bool


@dataclass(frozen=True)
class _ExcelRenderJob:
    source_path: Path
    source_alias: str
    sheet_name: str
    base_date: date
    effective_base_date: date
    frame_sum: pd.DataFrame
    frame_mean: pd.DataFrame


def parse_event_sheet_date(sheet_name: str) -> date | None:
    raw = sheet_name.strip()
    if raw.startswith(_RESPLIT_PREFIX):
        raw = raw.replace(_RESPLIT_PREFIX, "", 1)
    if not _DATE_SHEET_RE.fullmatch(raw):
        return None
    try:
        return datetime.strptime(raw, "%Y.%m.%d").date()
    except ValueError:
        return None


def collect_excel_event_candidates(input_excel: Path) -> list[ExcelEventCandidate]:
    if not input_excel.exists():
        raise ZipFlowError(f"入力Excelファイルが見つかりません: {input_excel}", exit_code=2)

    workbook = load_workbook(input_excel, data_only=True, read_only=True)
    try:
        candidates: list[ExcelEventCandidate] = []
        for sheet_name in workbook.sheetnames:
            event_date = parse_event_sheet_date(sheet_name)
            if event_date is None:
                continue
            candidates.append(
                ExcelEventCandidate(
                    event_date=event_date,
                    sheet_name=sheet_name,
                    is_resplit=sheet_name.startswith(_RESPLIT_PREFIX),
                )
            )
        candidates.sort(key=lambda item: (item.event_date, item.sheet_name))
        return candidates
    finally:
        workbook.close()


def export_excel_event_candidates_csv(
    *,
    input_excel: Path,
    output_all_csv: Path,
    output_unique_csv: Path,
) -> dict[str, object]:
    candidates = collect_excel_event_candidates(input_excel)
    output_all_csv.parent.mkdir(parents=True, exist_ok=True)
    output_unique_csv.parent.mkdir(parents=True, exist_ok=True)

    with output_all_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["event_date", "sheet_name", "is_resplit"])
        for item in candidates:
            writer.writerow([item.event_date.isoformat(), item.sheet_name, str(item.is_resplit).lower()])

    unique_dates = sorted({item.event_date for item in candidates})
    with output_unique_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["event_date"])
        for event_date in unique_dates:
            writer.writerow([event_date.isoformat()])

    return {
        "input_excel": str(input_excel),
        "output_all_csv": str(output_all_csv),
        "output_unique_csv": str(output_unique_csv),
        "candidate_count": len(candidates),
        "unique_date_count": len(unique_dates),
    }


def resolve_effective_base_date(base_date: date, graph_span: str) -> date:
    if graph_span == "3d_left":
        return base_date - timedelta(days=1)
    if graph_span == "3d_right":
        return base_date + timedelta(days=1)
    return base_date


def _to_datetime(value: object) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time(hour=0))
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ZipFlowError(
            "Excel時刻列(B列)の解釈に失敗しました。",
            exit_code=5,
        )
    return pd.Timestamp(parsed).to_pydatetime()


def _to_float(value: object) -> float:
    parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(parsed):
        raise ZipFlowError(
            "Excel雨量列(Q列)の数値変換に失敗しました。",
            exit_code=5,
        )
    return float(parsed)


def _excel_context_text(excel_path: Path, *, sheet_name: str, row_no: int | None = None) -> str:
    if row_no is None:
        return (
            f"対象ファイル: {excel_path.name}\n"
            f"ファイルパス: {excel_path}\n"
            f"対象シート: {sheet_name}"
        )
    return (
        f"対象ファイル: {excel_path.name}\n"
        f"ファイルパス: {excel_path}\n"
        f"対象シート: {sheet_name}\n"
        f"対象行: {row_no}"
    )


def _load_sheet_series(excel_path: Path, *, sheet_name: str, base_date: date) -> pd.DataFrame:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ZipFlowError(
                "指定シートが存在しません。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}",
                exit_code=5,
            )
        ws = wb[sheet_name]
        rows: list[dict[str, object]] = []
        for row_no, values in enumerate(ws.iter_rows(min_row=5, max_col=17, values_only=True), start=5):
            b = values[1] if len(values) > 1 else None
            q = values[16] if len(values) > 16 else None
            if b is None and q is None:
                continue
            if b is None or q is None:
                raise ZipFlowError(
                    "Excel時系列に欠損行があります（B列またはQ列が空欄）。\n"
                    f"{_excel_context_text(excel_path, sheet_name=sheet_name, row_no=row_no)}",
                    exit_code=5,
                )
            try:
                observed = _to_datetime(b)
            except ZipFlowError as exc:
                raise ZipFlowError(
                    f"{exc}\n{_excel_context_text(excel_path, sheet_name=sheet_name, row_no=row_no)}",
                    exit_code=exc.exit_code,
                ) from exc
            try:
                rainfall = _to_float(q)
            except ZipFlowError as exc:
                raise ZipFlowError(
                    f"{exc}\n{_excel_context_text(excel_path, sheet_name=sheet_name, row_no=row_no)}",
                    exit_code=exc.exit_code,
                ) from exc
            rows.append({"observed_at": observed, "rainfall_mm": rainfall})

        frame = pd.DataFrame(rows)
        if frame.empty:
            raise ZipFlowError(
                "Excel時系列が空です。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}",
                exit_code=5,
            )
        if len(frame) != 120:
            raise ZipFlowError(
                "Excel時系列点数が120ではありません。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}\n"
                f"期待点数: 120\n"
                f"実点数: {len(frame)}",
                exit_code=5,
            )

        if frame["observed_at"].duplicated().any():
            raise ZipFlowError(
                "Excel時刻列に重複があります。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}",
                exit_code=5,
            )
        if not frame["observed_at"].is_monotonic_increasing:
            raise ZipFlowError(
                "Excel時刻列が昇順ではありません。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}",
                exit_code=5,
            )

        diffs = frame["observed_at"].diff().dropna()
        if not diffs.eq(timedelta(hours=1)).all():
            raise ZipFlowError(
                "Excel時刻列が1時間間隔ではありません。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}",
                exit_code=5,
            )

        expected_start = datetime.combine(base_date - timedelta(days=2), time(hour=1))
        expected_end = datetime.combine(base_date + timedelta(days=3), time(hour=0))
        actual_start = pd.Timestamp(frame["observed_at"].iloc[0]).to_pydatetime()
        actual_end = pd.Timestamp(frame["observed_at"].iloc[-1]).to_pydatetime()
        if actual_start != expected_start or actual_end != expected_end:
            raise ZipFlowError(
                "Excel時系列の期間がシート日付と一致しません。\n"
                f"{_excel_context_text(excel_path, sheet_name=sheet_name)}\n"
                f"期待期間: {expected_start:%Y-%m-%d %H:%M} ～ {expected_end:%Y-%m-%d %H:%M}\n"
                f"実データ期間: {actual_start:%Y-%m-%d %H:%M} ～ {actual_end:%Y-%m-%d %H:%M}\n"
                "確認ポイント: シート名の日付 / B列の開始・終了時刻 / 入力ファイル選択",
                exit_code=5,
            )
        # Excel側の1時間は「01:00〜翌00:00」表記のため、グラフ系は0時起点へ正規化して扱う。
        frame["observed_at"] = frame["observed_at"] - timedelta(hours=1)
        return frame
    finally:
        wb.close()


def _build_excel_axis_tops(
    *,
    jobs: list[_ExcelRenderJob],
    render_span: str,
    ref_graph_kinds: tuple[str, ...],
    left_top_default: float,
    right_top_default: float,
) -> dict[tuple[str, str], tuple[float, float]]:
    span_days = 5 if render_span == "5d" else 3
    axis_tops: dict[tuple[str, str], tuple[float, float]] = {}
    for kind in ref_graph_kinds:
        left_max = 0.0
        right_max = 0.0
        for job in jobs:
            frame_src = job.frame_sum if kind == "sum" else job.frame_mean
            center = datetime.combine(job.effective_base_date, time(hour=0))
            start = center - timedelta(days=span_days // 2)
            end = start + timedelta(hours=(span_days * 24) - 1)
            span_frame = frame_src[(frame_src["observed_at"] >= start) & (frame_src["observed_at"] <= end)]
            window = prepare_reference_window(span_frame)
            left_max = max(left_max, float(window["rainfall_mm"].max()))
            right_max = max(right_max, float(window["cumulative_mm"].max()))
        axis_tops[(render_span, kind)] = compute_axis_tops(
            left_max=left_max,
            right_max=right_max,
            left_top_default=left_top_default,
            right_top_default=right_top_default,
        )
    return axis_tops


def run_excel_mode(config: ExcelRunConfig) -> dict[str, object]:
    if not config.input_excels:
        raise ZipFlowError("Excelモードでは入力Excelファイルを1件以上指定してください。", exit_code=2)
    missing_excels = [path for path in config.input_excels if not path.exists()]
    if missing_excels:
        raise ZipFlowError(f"入力Excelファイルが見つかりません: {missing_excels[0]}", exit_code=2)
    if not config.selected_events:
        raise ZipFlowError("Excelモードではイベント候補を1件以上選択してください。", exit_code=2)
    if config.graph_span not in ("5d", "3d_left", "3d_center", "3d_right"):
        raise ZipFlowError(f"未対応の graph_span です: {config.graph_span}", exit_code=2)
    if not config.ref_graph_kinds:
        raise ZipFlowError("グラフ指標を1つ以上選択してください。", exit_code=2)

    plot_ref_dir = config.output_root / "plots_reference"
    plot_ref_dir.mkdir(parents=True, exist_ok=True)
    log_dir = config.output_root / "logs"
    if config.enable_log:
        log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / "excel_mode.log"
    logger = build_logger(enable_file=config.enable_log, log_path=log_path)
    logger.info("Excelモード実行開始: inputs=%s", [str(p) for p in config.input_excels])
    logger.info(
        "選択イベント: %s",
        [f"[{e.source_alias}] {e.sheet_name}" for e in config.selected_events],
    )

    style_profile = None
    try:
        style_profile = load_style_profile(config.style_profile_path)
    except Exception as exc:  # noqa: BLE001
        raise ZipFlowError(f"スタイル読込に失敗しました: {exc}", exit_code=2) from exc

    render_span = "5d" if config.graph_span == "5d" else "3d"
    jobs: list[_ExcelRenderJob] = []
    for selected in config.selected_events:
        base_date = selected.event_date
        effective_base_date = resolve_effective_base_date(base_date, config.graph_span)
        frame_src = _load_sheet_series(selected.source_path, sheet_name=selected.sheet_name, base_date=base_date)
        observed_at = frame_src["observed_at"].tolist()
        rainfall = frame_src["rainfall_mm"].astype(float).tolist()
        frame_sum = build_metric_frame(observed_at=observed_at, weighted_sum=rainfall)
        frame_mean = build_metric_frame(observed_at=observed_at, weighted_sum=rainfall)
        logger.info(
            "シート検証OK: source=%s sheet=%s points=%s range=%s..%s",
            selected.source_alias,
            selected.sheet_name,
            len(frame_src),
            frame_src["observed_at"].min(),
            frame_src["observed_at"].max(),
        )
        jobs.append(
            _ExcelRenderJob(
                source_path=selected.source_path,
                source_alias=selected.source_alias,
                sheet_name=selected.sheet_name,
                base_date=base_date,
                effective_base_date=effective_base_date,
                frame_sum=frame_sum,
                frame_mean=frame_mean,
            )
        )

    axis_tops = _build_excel_axis_tops(
        jobs=jobs,
        render_span=render_span,
        ref_graph_kinds=config.ref_graph_kinds,
        left_top_default=style_profile.left_axis_top,
        right_top_default=style_profile.right_axis_top,
    )
    for kind in config.ref_graph_kinds:
        left_top, right_top = axis_tops[(render_span, kind)]
        logger.info("共通軸上限: span=%s kind=%s left_top=%.3f right_top=%.3f", render_span, kind, left_top, right_top)

    saved: list[Path] = []
    saved_png: list[Path] = []
    intermediate_jobs: list[dict[str, object]] = []
    for job in jobs:
        outputs = render_region_plots_reference(
            frame_sum=job.frame_sum,
            frame_mean=job.frame_mean,
            region_key=config.region_key,
            region_label=config.region_label,
            output_dir=plot_ref_dir,
            base_date=job.effective_base_date,
            graph_spans=(render_span,),
            ref_graph_kinds=config.ref_graph_kinds,
            export_svg=config.export_svg,
            on_conflict=config.on_conflict,
            style=style_profile,
            filename_prefix=build_excel_filename_prefix(job.source_alias),
            axis_tops=axis_tops,
        )
        saved.extend(outputs)
        saved_png.extend([path for path in outputs if path.suffix.lower() == ".png"])
        logger.info("グラフ出力完了: source=%s sheet=%s files=%s", job.source_alias, job.sheet_name, len(outputs))
        intermediate_jobs.append(
            {
                "base_date": job.base_date.strftime("%Y-%m-%d"),
                "reference_base_date": job.effective_base_date.strftime("%Y-%m-%d"),
                "region_key": config.region_key,
                "region_label": config.region_label,
                "graph_spans": [render_span],
                "ref_graph_kinds": list(config.ref_graph_kinds),
                "observed_at_jst": [str(ts) for ts in job.frame_sum["observed_at"].tolist()],
                "weighted_sum_mm": [float(v) for v in job.frame_sum["rainfall_mm"].tolist()],
                "weighted_mean_mm": [float(v) for v in job.frame_mean["rainfall_mm"].tolist()],
                "source_path": str(job.source_path),
                "source_alias": job.source_alias,
                "source_sheet_name": job.sheet_name,
            }
        )

    logger.info("Excelモード完了: plot_count=%s", len(saved))
    return {
        "base_dir": str(config.output_root),
        "plot_ref_dir": str(plot_ref_dir),
        "log_path": str(log_path) if config.enable_log else None,
        "zip_count": 0,
        "plot_count": len(saved),
        "csv_count": 0,
        "cell_csv_count": 0,
        "csv_readme_path": None,
        "event_count": len(config.selected_events),
        "intermediate_jobs": intermediate_jobs,
        "plot_ref_png_paths": [str(path) for path in saved_png],
    }
