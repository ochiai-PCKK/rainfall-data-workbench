from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from uc_rainfall_zipflow.errors import ZipFlowError
from uc_rainfall_zipflow.excel_application import (
    _load_sheet_series,
    collect_excel_event_candidates,
    parse_event_sheet_date,
    resolve_effective_base_date,
)
from uc_rainfall_zipflow.graph_builder import _extract_span_frame


def _write_event_workbook(path: Path, sheet_name: str, *, rows: int = 120) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    base_date = parse_event_sheet_date(sheet_name)
    assert base_date is not None
    start = datetime.combine(base_date - timedelta(days=2), datetime.min.time()) + timedelta(hours=1)
    for idx in range(rows):
        observed = start + timedelta(hours=idx)
        row = 5 + idx
        ws.cell(row=row, column=2).value = observed
        ws.cell(row=row, column=17).value = float(idx)
    wb.save(path)
    wb.close()


@pytest.mark.parametrize(
    ("sheet_name", "expected"),
    [
        ("2024.01.03", date(2024, 1, 3)),
        ("【再分割】2024.01.03", date(2024, 1, 3)),
        ("管理", None),
    ],
)
def test_parse_event_sheet_date(sheet_name: str, expected: date | None) -> None:
    assert parse_event_sheet_date(sheet_name) == expected


@pytest.mark.parametrize(
    ("graph_span", "expected"),
    [
        ("5d", date(2024, 1, 3)),
        ("3d_left", date(2024, 1, 2)),
        ("3d_center", date(2024, 1, 3)),
        ("3d_right", date(2024, 1, 4)),
    ],
)
def test_resolve_effective_base_date(graph_span: str, expected: date) -> None:
    assert resolve_effective_base_date(date(2024, 1, 3), graph_span) == expected


def test_collect_excel_event_candidates_sorts_candidates(tmp_path: Path) -> None:
    path = tmp_path / "events.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "2024.01.03"
    wb.create_sheet("【再分割】2024.01.01")
    wb.create_sheet("2023.12.31")
    wb.create_sheet("管理")
    wb.save(path)
    wb.close()

    candidates = collect_excel_event_candidates(path)

    assert [item.sheet_name for item in candidates] == [
        "2023.12.31",
        "【再分割】2024.01.01",
        "2024.01.03",
    ]
    assert [item.event_date for item in candidates] == [
        date(2023, 12, 31),
        date(2024, 1, 1),
        date(2024, 1, 3),
    ]


def test_load_sheet_series_normalizes_and_validates_window(tmp_path: Path) -> None:
    path = tmp_path / "series.xlsx"
    _write_event_workbook(path, "2024.01.03")

    frame = _load_sheet_series(path, sheet_name="2024.01.03", base_date=date(2024, 1, 3))

    assert len(frame) == 120
    assert frame["observed_at"].iloc[0] == datetime(2024, 1, 1, 0, 0)
    assert frame["observed_at"].iloc[-1] == datetime(2024, 1, 5, 23, 0)
    assert frame["rainfall_mm"].iloc[0] == 0.0
    assert frame["rainfall_mm"].iloc[-1] == 119.0


def test_load_sheet_series_rejects_non_120_points(tmp_path: Path) -> None:
    path = tmp_path / "series_bad.xlsx"
    _write_event_workbook(path, "2024.01.03", rows=119)

    with pytest.raises(ZipFlowError, match="期待点数: 120"):
        _load_sheet_series(path, sheet_name="2024.01.03", base_date=date(2024, 1, 3))


def test_extract_span_frame_uses_centered_window() -> None:
    base_date = date(2024, 1, 3)
    observed = pd.date_range("2024-01-01 00:00", periods=120, freq="h")
    frame = pd.DataFrame({"observed_at": observed, "weighted_sum": list(range(120))})

    span3 = _extract_span_frame(frame, base_date=base_date, span_days=3)
    span5 = _extract_span_frame(frame, base_date=base_date, span_days=5)

    assert len(span3) == 72
    assert span3["observed_at"].iloc[0] == pd.Timestamp("2024-01-02 00:00")
    assert span3["observed_at"].iloc[-1] == pd.Timestamp("2024-01-04 23:00")
    assert len(span5) == 120
    assert span5["observed_at"].iloc[0] == pd.Timestamp("2024-01-01 00:00")
    assert span5["observed_at"].iloc[-1] == pd.Timestamp("2024-01-05 23:00")
